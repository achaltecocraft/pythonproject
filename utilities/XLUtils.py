import openpyxl
from openpyxl import load_workbook

def getRowCount(file,sheetName):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    print("Row count is:",sheet_obj.max_row)
    return(sheet_obj.max_row)

def getColumCount(file,sheetName):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    print("Column Count is:",sheet_obj.max_column)
    return (sheet_obj.max_column)

def readData(file,sheetName,rownum,columnno):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    return sheet_obj.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(file)



#wb_obj = openpyxl.load_workbook(path)
#sheet_obj = wb_obj.active
# print the total number of rows
#print(sheet_obj.max_row)